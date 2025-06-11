from flask import Blueprint, render_template, request, redirect, url_for, flash, make_response, jsonify
from flask_login import login_required
from src.models import db, OrdemServico, Orcamento, Cliente, Material, NotaFiscal
from datetime import datetime, timedelta
from sqlalchemy import func, and_
import pandas as pd
from io import BytesIO
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill

relatorios_bp = Blueprint('relatorios', __name__)

@relatorios_bp.route('/')
@login_required
def index():
    return render_template('relatorios/index.html')

@relatorios_bp.route('/servicos')
@login_required
def servicos():
    # Parâmetros de filtro
    data_inicio = request.args.get('data_inicio')
    data_fim = request.args.get('data_fim')
    cliente_id = request.args.get('cliente_id')
    status = request.args.get('status')
    
    # Query base
    query = OrdemServico.query.join(Orcamento).join(Cliente)
    
    # Aplicar filtros
    if data_inicio:
        try:
            data_inicio_obj = datetime.strptime(data_inicio, '%Y-%m-%d')
            query = query.filter(OrdemServico.data_inicio >= data_inicio_obj)
        except ValueError:
            flash('Data de início inválida.', 'error')
    
    if data_fim:
        try:
            data_fim_obj = datetime.strptime(data_fim, '%Y-%m-%d')
            # Adicionar 23:59:59 para incluir todo o dia
            data_fim_obj = data_fim_obj.replace(hour=23, minute=59, second=59)
            query = query.filter(OrdemServico.data_inicio <= data_fim_obj)
        except ValueError:
            flash('Data de fim inválida.', 'error')
    
    if cliente_id:
        query = query.filter(Orcamento.cliente_id == cliente_id)
    
    if status:
        query = query.filter(OrdemServico.status == status)
    
    # Executar query
    ordens = query.order_by(OrdemServico.data_inicio.desc()).all()
    
    # Calcular estatísticas
    total_servicos = len(ordens)
    valor_total = sum(ordem.orcamento.valor_total for ordem in ordens)
    servicos_concluidos = len([o for o in ordens if o.status == 'concluido'])
    servicos_em_andamento = len([o for o in ordens if o.status == 'em_andamento'])
    servicos_cancelados = len([o for o in ordens if o.status == 'cancelado'])
    
    # Clientes para filtro
    clientes = Cliente.query.order_by(Cliente.nome).all()
    
    return render_template('relatorios/servicos.html',
                         ordens=ordens,
                         clientes=clientes,
                         filtros={
                             'data_inicio': data_inicio,
                             'data_fim': data_fim,
                             'cliente_id': cliente_id,
                             'status': status
                         },
                         estatisticas={
                             'total_servicos': total_servicos,
                             'valor_total': valor_total,
                             'servicos_concluidos': servicos_concluidos,
                             'servicos_em_andamento': servicos_em_andamento,
                             'servicos_cancelados': servicos_cancelados
                         })

@relatorios_bp.route('/servicos/excel')
@login_required
def servicos_excel():
    # Obter os mesmos filtros da página de relatórios
    data_inicio = request.args.get('data_inicio')
    data_fim = request.args.get('data_fim')
    cliente_id = request.args.get('cliente_id')
    status = request.args.get('status')
    
    # Query base
    query = OrdemServico.query.join(Orcamento).join(Cliente)
    
    # Aplicar filtros
    if data_inicio:
        try:
            data_inicio_obj = datetime.strptime(data_inicio, '%Y-%m-%d')
            query = query.filter(OrdemServico.data_inicio >= data_inicio_obj)
        except ValueError:
            pass
    
    if data_fim:
        try:
            data_fim_obj = datetime.strptime(data_fim, '%Y-%m-%d')
            data_fim_obj = data_fim_obj.replace(hour=23, minute=59, second=59)
            query = query.filter(OrdemServico.data_inicio <= data_fim_obj)
        except ValueError:
            pass
    
    if cliente_id:
        query = query.filter(Orcamento.cliente_id == cliente_id)
    
    if status:
        query = query.filter(OrdemServico.status == status)
    
    # Executar query
    ordens = query.order_by(OrdemServico.data_inicio.desc()).all()
    
    # Criar workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Relatório de Serviços"
    
    # Estilos
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    # Cabeçalho
    headers = [
        "Ordem de Serviço", "Cliente", "Orçamento", "Data Início", 
        "Data Conclusão", "Status", "Descrição", "Valor Total"
    ]
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
    
    # Dados
    for row, ordem in enumerate(ordens, 2):
        ws.cell(row=row, column=1, value=ordem.numero_ordem)
        ws.cell(row=row, column=2, value=ordem.orcamento.cliente.nome)
        ws.cell(row=row, column=3, value=ordem.orcamento.numero_orcamento)
        ws.cell(row=row, column=4, value=ordem.data_inicio.strftime('%d/%m/%Y'))
        ws.cell(row=row, column=5, value=ordem.data_conclusao.strftime('%d/%m/%Y') if ordem.data_conclusao else 'Em andamento')
        ws.cell(row=row, column=6, value=ordem.status.replace('_', ' ').title())
        ws.cell(row=row, column=7, value=ordem.orcamento.descricao_servico)
        ws.cell(row=row, column=8, value=f"R$ {ordem.orcamento.valor_total:.2f}")
    
    # Ajustar largura das colunas
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Salvar em buffer
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    # Criar resposta
    response = make_response(buffer.getvalue())
    response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    response.headers['Content-Disposition'] = f'attachment; filename=relatorio_servicos_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    
    return response

@relatorios_bp.route('/servicos/pdf')
@login_required
def servicos_pdf():
    # Obter os mesmos filtros da página de relatórios
    data_inicio = request.args.get('data_inicio')
    data_fim = request.args.get('data_fim')
    cliente_id = request.args.get('cliente_id')
    status = request.args.get('status')
    
    # Query base
    query = OrdemServico.query.join(Orcamento).join(Cliente)
    
    # Aplicar filtros
    if data_inicio:
        try:
            data_inicio_obj = datetime.strptime(data_inicio, '%Y-%m-%d')
            query = query.filter(OrdemServico.data_inicio >= data_inicio_obj)
        except ValueError:
            pass
    
    if data_fim:
        try:
            data_fim_obj = datetime.strptime(data_fim, '%Y-%m-%d')
            data_fim_obj = data_fim_obj.replace(hour=23, minute=59, second=59)
            query = query.filter(OrdemServico.data_inicio <= data_fim_obj)
        except ValueError:
            pass
    
    if cliente_id:
        query = query.filter(Orcamento.cliente_id == cliente_id)
    
    if status:
        query = query.filter(OrdemServico.status == status)
    
    # Executar query
    ordens = query.order_by(OrdemServico.data_inicio.desc()).all()
    
    try:
        from src.utils.pdf_generator import gerar_pdf_relatorio_servicos
        pdf_content = gerar_pdf_relatorio_servicos(ordens, {
            'data_inicio': data_inicio,
            'data_fim': data_fim,
            'cliente_id': cliente_id,
            'status': status
        })
        
        response = make_response(pdf_content)
        response.headers['Content-Type'] = 'application/pdf'
        response.headers['Content-Disposition'] = f'attachment; filename=relatorio_servicos_{datetime.now().strftime("%Y%m%d_%H%M%S")}.pdf'
        
        return response
    except Exception as e:
        flash('Erro ao gerar PDF do relatório.', 'error')
        return redirect(url_for('relatorios.servicos'))

@relatorios_bp.route('/dashboard')
@login_required
def dashboard():
    # Período padrão: últimos 30 dias
    data_fim = datetime.now()
    data_inicio = data_fim - timedelta(days=30)
    
    # Estatísticas gerais
    total_clientes = Cliente.query.count()
    total_orcamentos = Orcamento.query.count()
    total_ordens = OrdemServico.query.count()
    total_notas_fiscais = NotaFiscal.query.count()
    
    # Ordens por status
    ordens_em_andamento = OrdemServico.query.filter_by(status='em_andamento').count()
    ordens_concluidas = OrdemServico.query.filter_by(status='concluido').count()
    ordens_canceladas = OrdemServico.query.filter_by(status='cancelado').count()
    
    # Faturamento do mês
    faturamento_mes = db.session.query(func.sum(Orcamento.valor_total)).join(OrdemServico).filter(
        and_(
            OrdemServico.status == 'concluido',
            OrdemServico.data_conclusao >= data_inicio,
            OrdemServico.data_conclusao <= data_fim
        )
    ).scalar() or 0
    
    # Orçamentos pendentes
    orcamentos_pendentes = Orcamento.query.filter_by(status='pendente').count()
    
    # Materiais com estoque baixo
    materiais_estoque_baixo = Material.query.filter(Material.estoque <= Material.estoque_minimo).count()
    
    return render_template('relatorios/dashboard.html',
                         estatisticas={
                             'total_clientes': total_clientes,
                             'total_orcamentos': total_orcamentos,
                             'total_ordens': total_ordens,
                             'total_notas_fiscais': total_notas_fiscais,
                             'ordens_em_andamento': ordens_em_andamento,
                             'ordens_concluidas': ordens_concluidas,
                             'ordens_canceladas': ordens_canceladas,
                             'faturamento_mes': faturamento_mes,
                             'orcamentos_pendentes': orcamentos_pendentes,
                             'materiais_estoque_baixo': materiais_estoque_baixo
                         })

